# !/usr/bin/env python3
# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------------
# Created By            : xxxxxxx
# Email                 : xxxxxxx
# Created Date          : 18/04/2024
# Last Modified Date    : 18/04/2024
# version               : '1.0'
# Description           : This tool will extract the APK features and detect for obfuscation,
#                         obfuscation tool and the technique
# Usage                 :Large-Scale Investigation in Google Play Store
# ---------------------------------------------------------------------------

import os
import re
import csv
import torch
import joblib
import logging

import pandas as pd
import torch.nn as nn
import numpy as np
import torch.optim as optim
import torch.nn.functional as F

from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl import Workbook
from torch_geometric.data import Data
from androguard.misc import AnalyzeAPK
from androguard.core.bytecodes.apk import APK
from androguard.core.bytecodes.dvm import DalvikVMFormat

print("*********************************************************************************************************")
print("******************************* APK Level Analyser - Version 1.0 -2023 APKs *****************************")
print("*********************************************************************************************************")

INPUT_DIM = 37
HIDDEN_DIM = 32
OUTPUT_DIM = 1
OBFUSCATION_DETECTION_MODEL_LEARNING_RATE = 0.001

WORKING_DIR = "Add your working directory"
OBFUSCATION_MODEL_NAME = "Obfuscation Detector Model Name"
PROGUARD_MODEL_NAME = "ProGuard Model Name"
DASHO_MODEL_NAME = "DashO Model Name"
ALLATORI_MODEL_NAME = "Allatori Model Name"
IR_MODEL_NAME = "IR Model Name"
CF_MODEL_NAME = "CF Model Name"
SE_MODEL_NAME = "SE Model Name"

LOG_FILE = WORKING_DIR + "/analyser_log.log"
LOGGING_LEVEL = logging.INFO
EXCEL_FILE = WORKING_DIR + "/investigation_results_APK_files3.xlsx"

APK_FILE = "example_apk_file.csv"

logger = logging.getLogger(__name__)
logging.basicConfig(filename=LOG_FILE, encoding='utf-8', level=LOGGING_LEVEL,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def is_readable_string(s):
    # Use a regular expression to check if the string contains only printable characters
    return bool(re.match(r'^[\x20-\x7E]+$', s))


def remove_unreadable_strings(strings):
    try:
        cleaned_strings = [s for s in strings if is_readable_string(s)]
        return cleaned_strings
    except Exception as e:
        logger.error("Error occurred while removing the unreadable strings: {}".format(e))


def calculate_ins_occ(d):
    total = 0
    goto = 0
    invoke = 0
    nop = 0
    if_ins = 0
    move = 0
    # Extract opcodes from each method
    for c in d.get_classes():
        # print(c)
        for m in c.get_methods():
            m.get_name()
            for i in m.get_instructions():
                inst = i.get_name()
                total += 1
                if 'goto' in inst:
                    goto += 1
                elif 'invoke' in inst:
                    invoke += 1
                elif 'nop' in inst:
                    nop += 1
                elif 'if' in inst:
                    if_ins += 1
                elif 'move' in inst:
                    move += 1
    return goto, invoke, nop, if_ins, move, total


#############################################################################
#############################################################################

def extract_features_from_dex(apk):
    strings = []
    class_names = []
    method_names = []
    field_names = []
    nop_count = 0
    invoke_count = 0
    move_count = 0
    if_count = 0
    goto_count = 0
    total_ins = 0
    dex_files = None
    logger.info("Extracting Instructions, Identifier Names and Other Strings from MULTI/SINGLE Dex file")
    try:
        apk_parser = APK(apk)
        dex_files = apk_parser.get_all_dex()
    except Exception as e:
        logger.error("Error occurred while extracting Dex File: {}".format(e))

    for dex in dex_files:
        dalvik = DalvikVMFormat(dex)
        tmp_goto, tmp_invoke, tmp_nop, tmp_if_ins, tmp_move, tmp_total = calculate_ins_occ(dalvik)
        nop_count += tmp_nop
        invoke_count += tmp_invoke
        move_count += tmp_move
        if_count += tmp_if_ins
        goto_count += tmp_goto
        total_ins += tmp_total
        for c in dalvik.get_classes():
            class_names.append(os.path.basename(c.get_name()).split('/')[-1].replace(";", ""))
            for m in c.get_methods():
                method_names.append(m.get_name())
            for f in c.get_fields():
                field_names.append(f.get_name())
        for st in dalvik.get_strings():
            strings.append(st)
    # Need to remove unwanted method names from List
    # Remove <init> and <cinit> from method list
    while "<init>" in method_names:
        method_names.remove("<init>")
    while "<cinit>" in method_names:
        method_names.remove("<cinit>")

    # Remove unreadable strings from the string list
    strings_cleared = remove_unreadable_strings(strings)
    # Convert the filter lists to sets for faster membership testing
    class_names_set = set(class_names)
    method_names_set = set(method_names)
    field_names_set = set(field_names)

    # Use a single list comprehension to filter words
    filtered_words = [word for word in strings_cleared if
                      word not in class_names_set and word not in method_names_set and word not in field_names_set]
    return [nop_count,
            invoke_count,
            move_count,
            if_count,
            goto_count,
            total_ins,
            class_names,
            method_names,
            field_names,
            filtered_words]


def calculate_avg_occ(list_of_identifiers, average_constant):
    # Regular expression patterns to match strings with special characters and numeric characters
    special_char_pattern = r'\w*[^\w\s]+\w*'
    numeric_pattern = r'\d'

    feature_list = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
    word_lengths = []
    total_word_count = len(list_of_identifiers)
    # print("Total String Count: {}".format(total_word_count))

    # Iterate through the list of strings
    for string in list_of_identifiers:
        word_lengths.append(len(string))
        # print("String - {} | Length - {}".format(string,len(string)))
        if re.search(special_char_pattern, string) and re.search(numeric_pattern, string):
            feature_list[0] += average_constant / total_word_count
            # print("Special + Numeric")
        elif re.search(numeric_pattern, string):
            feature_list[1] += average_constant / total_word_count
            # print("Numeric")
        elif re.search(special_char_pattern, string):
            feature_list[2] += average_constant / total_word_count
            # print("Special")

    for string in list_of_identifiers:
        length = len(string)
        if length == 1:
            feature_list[3] += average_constant / total_word_count
        elif length == 2:
            feature_list[4] += average_constant / total_word_count
        elif length == 3:
            feature_list[5] += average_constant / total_word_count
        elif length == 4:
            feature_list[6] += average_constant / total_word_count
        else:
            feature_list[7] += average_constant / total_word_count

    return feature_list


class MLP(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLP, self).__init__()
        self.lin1 = nn.Linear(input_dim, hidden_dim)  # max pool 32, mean pool 32, apk level features 32
        self.lin2 = nn.Linear(hidden_dim, hidden_dim)
        self.lin3 = nn.Linear(hidden_dim, hidden_dim)
        self.lin4 = nn.Linear(hidden_dim, output_dim)

    def forward(self, data):
        x, edge_index, batch = data.x, data.edge_index, data.batch

        # Basic MLP Layer
        x = F.relu(self.lin1(x))
        x = F.relu(self.lin2(x))
        x = F.relu(self.lin3(x))
        x = self.lin4(x)
        return x


def load_mlp_model(learning_rate, model_path):
    # Loading Models
    try:
        model_mlp = MLP(INPUT_DIM, HIDDEN_DIM, OUTPUT_DIM)
        optimizer_mlp = optim.Adam(model_mlp.parameters(), lr=learning_rate)
        checkpoint = torch.load(model_path)
        model_mlp.load_state_dict(checkpoint['model_state_dict'])
        optimizer_mlp.load_state_dict(checkpoint['optimizer_state_dict'])
        return model_mlp
    except Exception as e:
        logger.error("Error occurred while loading the MLP Model: {}".format(e))


def generate_features(apk):
    feature_list_obfuscator = []
    feature_list = []
    logger.info("Creating Feature Vector for: {}".format(apk))
    extracted_feature_list = None
    try:
        logger.info("Trying to extract features from APK")
        extracted_feature_list = extract_features_from_dex(apk)

    except Exception as e:
        logger.error("Error occurred while extracting the features")

    if extracted_feature_list is not None:
        feature_list.append(extracted_feature_list[0] * 100 / extracted_feature_list[5])
        feature_list.append(extracted_feature_list[1] * 100 / extracted_feature_list[5])
        feature_list.append(extracted_feature_list[2] * 100 / extracted_feature_list[5])
        feature_list.append(extracted_feature_list[3] * 100 / extracted_feature_list[5])
        feature_list.append(extracted_feature_list[4] * 100 / extracted_feature_list[5])

        # Calculate Features for Classes, Methods, and Fields
        feature_list.extend(calculate_avg_occ(extracted_feature_list[6], 100))
        feature_list.extend(calculate_avg_occ(extracted_feature_list[7], 100))
        feature_list.extend(calculate_avg_occ(extracted_feature_list[8], 100))
        feature_list.extend(calculate_avg_occ(extracted_feature_list[9], 100))

        feature_list_obfuscator.append(extracted_feature_list[0] * 1 / extracted_feature_list[5])
        feature_list_obfuscator.append(extracted_feature_list[1] * 1 / extracted_feature_list[5])
        feature_list_obfuscator.append(extracted_feature_list[2] * 1 / extracted_feature_list[5])
        feature_list_obfuscator.append(extracted_feature_list[3] * 1 / extracted_feature_list[5])
        feature_list_obfuscator.append(extracted_feature_list[4] * 1 / extracted_feature_list[5])

        # Calculate Features for Classes, Methods, and Fields
        feature_list_obfuscator.extend(calculate_avg_occ(extracted_feature_list[6], 1))
        feature_list_obfuscator.extend(calculate_avg_occ(extracted_feature_list[7], 1))
        feature_list_obfuscator.extend(calculate_avg_occ(extracted_feature_list[8], 1))
        feature_list_obfuscator.extend(calculate_avg_occ(extracted_feature_list[9], 1))

        apk_level_features = torch.tensor(feature_list_obfuscator)
        apk_level_feature = Data(x=apk_level_features)

        # print("Features: {}".format(feature_list))
        return apk_level_feature, feature_list
    else:
        logger.error("Error occurred while extracting the features: Detected as None")
        return None, None


# APK Name, Year, Obfuscated?, Tool1, Probability, Tool2, Probability, Tool3, Probability Technique1, Probability
if __name__ == "__main__":
    logger.info("Obfuscation Analyser Started")

    obfuscation_detection_model_file = WORKING_DIR + '/BestModels/' + OBFUSCATION_MODEL_NAME

    proguard_model_file = WORKING_DIR + '/BestModels/' + PROGUARD_MODEL_NAME
    dasho_model_file = WORKING_DIR + '/BestModels/' + DASHO_MODEL_NAME
    allatori_model_file = WORKING_DIR + '/BestModels/' + ALLATORI_MODEL_NAME

    ir_model_file = WORKING_DIR + '/BestModels/' + IR_MODEL_NAME
    cf_model_file = WORKING_DIR + '/BestModels/' + CF_MODEL_NAME
    se_model_file = WORKING_DIR + '/BestModels/' + SE_MODEL_NAME

    # Path to your CSV file
    csv_file = WORKING_DIR + '/apk_list/' + APK_FILE

    excel_sheet_row_number = 0
    apk_number = 0

    ###### LOADING PREDICTION MODELS #######
    # Load Main Obfuscation Detection Model
    obfuscation_detection_model = load_mlp_model(
        learning_rate=OBFUSCATION_DETECTION_MODEL_LEARNING_RATE,
        model_path=obfuscation_detection_model_file)

    # Load Tool Detection Models
    proguard_detection_model = None
    dasho_detection_model = None
    allatori_detection_model = None
    proguard_prediction, proguard_probability = None, None
    dasho_probability, dasho_prediction = None, None
    allatori_probability, allatori_prediction = None, None
    ir_prediction, ir_probability = None, None
    cf_prediction, cf_probability = None, None
    se_prediction, se_probability = None, None

    try:
        proguard_detection_model = joblib.load(proguard_model_file)
        dasho_detection_model = joblib.load(dasho_model_file)
        allatori_detection_model = joblib.load(allatori_model_file)
        logger.info("Loading completed: Tool Detection Models")
    except Exception as e:
        logger.error("Error occurred while loading the Tool Model: {}".format(e))

    # Load Technique Detection Models
    ir_detection_model, cf_detection_model, se_detection_model = None, None, None
    try:
        ir_detection_model = joblib.load(ir_model_file)
        cf_detection_model = joblib.load(cf_model_file)
        se_detection_model = joblib.load(se_model_file)
        logger.info("Loading completed: Technique Detection Models")
    except Exception as e:
        logger.error("Error occurred while loading the Technique Model: {}".format(e))
    ###### DONE LOADING PREDICTION MODELS #######

    # Check if the Excel file already exists
    if os.path.exists(EXCEL_FILE):
        # Load the existing workbook
        workbook = load_workbook(EXCEL_FILE)
        # Select the active worksheet
        worksheet = workbook.active
    else:
        # Create a new workbook and worksheet if the file doesn't exist
        workbook = Workbook()
        worksheet = workbook.active
        # Write the header row to the Excel file
        header_values = ['App_ID', 'Last_Updated', 'App_Genre',
                         'Is_Obfuscated', 'Obfuscate_Probability',
                         'Assigned_Tool',
                         'ProGuard_Prediction', 'ProGuard_Probability',
                         'DashO_Prediction', 'DashO_Probability',
                         'Allatori_Prediction', 'Allatori_Probability',
                         'IR_Prediction', 'IR_Probability',
                         'CF_Prediction', 'CF_Probability',
                         'SE_Prediction', 'SE_Probability']
        worksheet.append(header_values)

    # Initialize the row number in the Excel sheet
    excel_sheet_row_number = worksheet.max_row + 1

    # Open the CSV file for reading
    with open(csv_file, 'r', newline='') as file:
        # Create a CSV reader object
        excel_list = []
        reader = csv.reader(file)
        # Skip the header row
        next(reader)
        reader = tqdm(reader, desc="Processing APKs", unit=" APK")
        for row in reader:
            apk_number += 1
            ####### Read The Row and create the APK File Path ###########
            app_id = row[0]
            last_updated = row[1]
            app_genre = row[2]
            folder_id = 0

            # Create Server Path
            server_path = "Create your APK file path here......."

            if os.path.exists(server_path):
                print("\nAnalysing APK: {}............".format(app_id))
                logger.info("Analysing APK: {}".format(app_id))

                logger.debug("Path: {}".format(server_path))

                obfuscation_feature, tool_and_technique_feature = generate_features(server_path)
                obfuscation_prediction = False
                if obfuscation_feature is not None:
                    obfuscation_output = obfuscation_detection_model(obfuscation_feature)
                    obfuscation_probability = torch.sigmoid(obfuscation_output.squeeze())
                    obfuscation_prediction = (obfuscation_probability > 0.5).float()

                    initial_detection_results = [app_id, last_updated, app_genre,
                                                 obfuscation_prediction.item(), obfuscation_probability.item()]
                    excel_list = initial_detection_results + [''] * 13
                else:
                    logger.error("Extracted Features are None. Cannot Analyse Further.")
                    initial_detection_results = [app_id, last_updated, folder_id, "Data_Error"]
                    excel_list = initial_detection_results + [''] * 14
                    obfuscation_prediction = False

                if obfuscation_prediction:
                    header = ["avg_nop", "avg_invoke", "avg_move", "avg_if", "avg_goto", "avg_class_num_char",
                              "avg_class_num", "avg_class_char", "avg_class_l1", "avg_class_l2", "avg_class_l3",
                              "avg_class_l4", "avg_class_ln", "avg_method_num_char", "avg_method_num",
                              "avg_method_char",
                              "avg_method_l1", "avg_method_l2", "avg_method_l3", "avg_method_l4", "avg_method_ln",
                              "avg_field_num_char", "avg_field_num", "avg_field_char", "avg_field_l1", "avg_field_l2",
                              "avg_field_l3", "avg_field_l4", "avg_field_ln", "avg_string_num_char", "avg_string_num",
                              "avg_string_char", "avg_string_l1", "avg_string_l2", "avg_string_l3", "avg_string_l4",
                              "avg_string_ln"]
                    # Convert feature to DF
                    feature_df = pd.DataFrame([tool_and_technique_feature], columns=header)

                    # ProGuard
                    if proguard_detection_model is not None:
                        proguard_probability = proguard_detection_model.predict_proba(feature_df)
                        proguard_prediction = proguard_detection_model.predict(feature_df)
                    else:
                        logger.error("ProGuard Detection Model is Not-Found")

                    # DashO
                    if dasho_detection_model is not None:
                        dasho_probability = dasho_detection_model.predict_proba(feature_df)
                        dasho_prediction = dasho_detection_model.predict(feature_df)
                    else:
                        logger.error("DashO Detection Model is Not-Found")

                    # Allatori
                    if allatori_detection_model is not None:
                        allatori_probability = allatori_detection_model.predict_proba(feature_df)
                        allatori_prediction = allatori_detection_model.predict(feature_df)
                    else:
                        logger.error("Allatori Detection Model is Not-Found")

                    probabilities = np.array(
                        [proguard_probability[0, 1], dasho_probability[0, 1], allatori_probability[0, 1]])

                    # Define the class labels
                    class_labels = ['ProGuard', 'DashO', 'Allatori']

                    # Check if all probabilities are less than 0.5
                    if np.all(probabilities < 0.5):
                        assigned_class = 'Other'
                    else:
                        # Get the index of the maximum probability
                        max_index = np.argmax(probabilities)
                        # Assign the class based on the maximum probability
                        assigned_class = class_labels[max_index]

                    if ir_detection_model is not None:
                        ir_probability = ir_detection_model.predict_proba(feature_df)
                        ir_prediction = ir_detection_model.predict(feature_df)
                    else:
                        logger.error("IR Detection Model is Not-Found")

                    if cf_detection_model is not None:
                        cf_probability = cf_detection_model.predict_proba(feature_df)
                        cf_prediction = cf_detection_model.predict(feature_df)
                    else:
                        logger.error("CF Detection Model is Not-Found")

                    if se_detection_model is not None:
                        se_probability = se_detection_model.predict_proba(feature_df)
                        se_prediction = se_detection_model.predict(feature_df)
                    else:
                        logger.error("SE Detection Model is Not-Found")

                    excel_list[5:] = [assigned_class,
                                      proguard_prediction[0], proguard_probability[0, 1],
                                      dasho_prediction[0], dasho_probability[0, 1],
                                      allatori_prediction[0], allatori_probability[0, 1],
                                      ir_prediction[0], ir_probability[0, 1],
                                      cf_prediction[0], cf_probability[0, 1],
                                      se_prediction[0], se_probability[0, 1]]
                    logger.info(
                        "Prediction Results: {}".format(app_id, last_updated, obfuscation_prediction,
                                                        obfuscation_probability))
                    logger.info("Prediction Results: {}".format(excel_list))
            else:
                excel_list = [app_id, last_updated, folder_id, "Not_Found"]
                print("\nAPK: {} Does not Exist in Folder: {}".format(app_id, folder_id))
                logger.error("APK: {} Does not exist in Folder: {}".format(app_id, folder_id))

            worksheet.append(excel_list)
            workbook.save(EXCEL_FILE)
            excel_sheet_row_number += 1
    print("Done Analysing APKs.....")
    logger.info("Done Analysing APKs....")
